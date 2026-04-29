import gdown
import os

DB_PATH = "apmi_pms.db"

def download_db_if_needed():
    if not os.path.exists(DB_PATH):
        print("📥 Downloading database from Google Drive...")
        try:
            gdown.download(
                f"https://drive.google.com/uc?id=1J1_EbyJLpAikKTLjP9aFEMp7zGjO3G6s",
                DB_PATH, quiet=False
            )
            print("✅ Database downloaded")
        except Exception as e:
            print(f"❌ Download failed: {e}")
    else:
        print("✅ Database found locally")

download_db_if_needed()

import gdown
import os


def download_db_if_needed():
    """Downloads apmi_pms.db from Google Drive if not present locally."""
    if not os.path.exists(DB_PATH):
        print("📥 Database not found locally — downloading from Google Drive...")
        try:
            gdown.download(
                "https://drive.google.com/uc?id=1J1_EbyJLpAikKTLjP9aFEMp7zGjO3G6s://drive.google.com/drive/folders/1oVpilE9DTfK7jtA1SdVol_T6XXrcwtOo?usp=sharing",
                DB_PATH, quiet=False
            )
            print("✅ Database downloaded successfully")
        except Exception as e:
            print(f"❌ Download failed: {e}")
            print("Make sure the file is shared as \'Anyone with the link can view\'")
    else:
        print(f"✅ Database found locally: {DB_PATH}")

download_db_if_needed()

import streamlit as st
import pandas as pd
import numpy as np
import sqlite3
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
import os
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")

st.set_page_config(
    page_title="APMI PMS Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
.stApp { background-color: #f8fafc; }
.metric-card {
    background: white; border-radius: 12px; padding: 20px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1); border-left: 4px solid #2563eb; margin-bottom: 12px;
}
.section-header {
    background: linear-gradient(135deg, #1e40af, #3b82f6); color: white;
    padding: 12px 20px; border-radius: 8px; margin: 20px 0 16px 0;
    font-size: 16px; font-weight: 600;
}
.info-box {
    background: #eff6ff; border: 1px solid #bfdbfe; border-radius: 8px;
    padding: 12px 16px; margin: 8px 0; font-size: 14px; color: #1e40af;
}
.warn-box {
    background: #fffbeb; border: 1px solid #fde68a; border-radius: 8px;
    padding: 12px 16px; margin: 8px 0; font-size: 14px; color: #92400e;
}
.error-box {
    background: #fef2f2; border: 1px solid #fecaca; border-radius: 8px;
    padding: 12px 16px; margin: 8px 0; font-size: 14px; color: #991b1b;
}
footer { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


COLORS = {
    "Equity": "#2563eb", "Debt": "#7c3aed", "Hybrid": "#059669",
    "Multi Asset": "#d97706", "Discretionary": "#2563eb",
    "Non-Discretionary": "#7c3aed", "Inflow": "#16a34a",
    "Outflow": "#dc2626", "Neutral": "#94a3b8", "benchmark": "#f59e0b",
}

RETURN_COLS = {
    "return_1m": "1 Month", "return_3m": "3 Months", "return_6m": "6 Months",
    "return_1y": "1 Year",  "return_2y": "2 Years",  "return_3y": "3 Years",
    "return_4y": "4 Years", "return_5y": "5 Years",  "return_si": "Since Inception",
}

PERIOD_TO_BENCH = {
    "return_1m": ("bench_return_1m", "alpha_1m"),
    "return_3m": ("bench_return_3m", "alpha_3m"),
    "return_6m": ("bench_return_6m", "alpha_6m"),
    "return_1y": ("bench_return_1y", "alpha_1y"),
    "return_3y": ("bench_return_3y", "alpha_3y"),
    "return_5y": ("bench_return_5y", "alpha_5y"),
}

FLAG_LABELS = {
    "CLEAN":                "✅ Clean",
    "STRUCTURED_PRODUCT":   "🧾 Structured Product",
    "NEGATIVE_AUM_ERROR":   "❌ Negative AUM (APMI Error)",
    "WOUND_DOWN":           "⚰️ Wound Down",
    "APMI_REPORTED_ANOMALY":"⚠️ APMI Reported Anomaly",
    "SUSPECTED_DATA_ERROR": "🚨 Suspected Data Error",
}

BASE_EXPORT_COLS = ["snapshot_date", "pms_provider", "ia_name", "strategy_type",
                    "service_type", "aum_cr", "data_quality_flag"]


# ─────────────────────────────────────────────────────────
# EXCEL MULTI-SHEET DOWNLOAD
# ─────────────────────────────────────────────────────────
def build_excel_all_periods(df_source, base_cols, df_bench=None):
    from openpyxl.styles import PatternFill, Font, Alignment
    output = io.BytesIO()

    summary_keep = [c for c in base_cols if c in df_source.columns]
    all_ret      = [c for c in RETURN_COLS if c in df_source.columns]
    df_summary   = df_source[summary_keep + all_ret].copy()
    df_summary   = df_summary.rename(columns=dict(RETURN_COLS))
    df_summary   = df_summary.rename(columns={
        "pms_provider": "Provider", "ia_name": "Investment Approach",
        "strategy_type": "Strategy", "service_type": "Service Type",
        "aum_cr": "AUM (Cr)", "snapshot_date": "Snapshot Date",
        "data_quality_flag": "Flag",
    })

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    green_fill  = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    red_fill    = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="All Periods Summary", index=False)
        ws = writer.sheets["All Periods Summary"]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes = "A2"
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        for ret_col, period_label in RETURN_COLS.items():
            if ret_col not in df_source.columns:
                continue
            keep = [c for c in base_cols if c in df_source.columns]
            df_p = df_source[keep + [ret_col]].copy()

            bench_col, alpha_col = PERIOD_TO_BENCH.get(ret_col, (None, None))
            if bench_col and bench_col in df_source.columns:
                df_p["Nifty 50 TRI"] = df_source[bench_col].values
            if alpha_col and alpha_col in df_source.columns:
                df_p["Alpha vs Nifty"] = df_source[alpha_col].values

            df_p = df_p.rename(columns={
                "pms_provider": "Provider", "ia_name": "Investment Approach",
                "strategy_type": "Strategy", "service_type": "Service Type",
                "aum_cr": "AUM (Cr)", "snapshot_date": "Snapshot Date",
                "data_quality_flag": "Flag", ret_col: f"Return ({period_label})",
            })
            ret_display = f"Return ({period_label})"
            if ret_display in df_p.columns:
                df_p = df_p.sort_values(ret_display, ascending=False)

            sheet_name = period_label[:31]
            df_p.to_excel(writer, sheet_name=sheet_name, index=False)
            ws2 = writer.sheets[sheet_name]
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws2.freeze_panes = "A2"

            ret_idx = next((i for i, c in enumerate(ws2[1], 1) if c.value == ret_display), None)
            if ret_idx:
                for row in ws2.iter_rows(min_row=2, min_col=ret_idx, max_col=ret_idx):
                    for cell in row:
                        try:
                            v = float(cell.value)
                            cell.fill = green_fill if v > 0 else (red_fill if v < 0 else PatternFill())
                        except (TypeError, ValueError):
                            pass

            for col in ws2.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output.seek(0)
    return output.getvalue()


def excel_button(df_source, base_cols, filename, df_bench=None,
                 label="📥 Excel — All Periods (9 sheets)"):
    if df_source.empty:
        return
    try:
        data = build_excel_all_periods(df_source, base_cols, df_bench=df_bench)
        st.download_button(
            label=label, data=data, file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="One sheet per return period (1M 3M 6M 1Y 2Y 3Y 4Y 5Y SI) + Summary sheet with all periods side-by-side. Return column is green for positive, red for negative."
        )
    except Exception as e:
        st.caption(f"Excel export unavailable: {e}")


def csv_button(df, filename, label="⬇️ CSV"):
    st.download_button(label, df.to_csv(index=False), filename, "text/csv")


# ─────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────
@st.cache_data(ttl=3600)
def load_table(table):
    if not os.path.exists(DB_PATH):
        return pd.DataFrame()
    try:
        conn = sqlite3.connect(DB_PATH)
        df   = pd.read_sql(f"SELECT * FROM {table}", conn)
        conn.close()
        return df
    except:
        return pd.DataFrame()

def load_performance():
    df = load_table("raw_performance")
    if not df.empty:
        df["snapshot_date"] = pd.to_datetime(df["snapshot_date"])
        df = df.loc[:, ~df.columns.duplicated()]
        if "data_quality_flag" in df.columns:
            df["data_quality_flag"] = df["data_quality_flag"].fillna("CLEAN")
    return df

def load_benchmarks():
    df = load_table("raw_benchmarks")
    if not df.empty:
        df["snapshot_date"] = pd.to_datetime(df["snapshot_date"])
    return df

def load_risk():
    return load_table("risk_metrics")

def load_aum_decomp():
    df = load_table("aum_decomposition")
    if not df.empty:
        df["snapshot_date"] = pd.to_datetime(df["snapshot_date"])
    return df

def load_turnover():
    df = load_table("ia_turnover")
    if not df.empty:
        df["snapshot_date"] = pd.to_datetime(df["snapshot_date"])
    return df

@st.cache_data(ttl=3600)
def load_data_notes():
    if not os.path.exists(DB_PATH):
        return pd.DataFrame()
    try:
        conn = sqlite3.connect(DB_PATH)
        df   = pd.read_sql("SELECT * FROM data_notes", conn)
        conn.close()
        return df
    except:
        return pd.DataFrame()

def load_performance_with_benchmarks():
    df_perf  = load_performance()
    df_bench = load_benchmarks()
    if df_bench.empty or df_perf.empty:
        return df_perf
    nifty = df_bench[df_bench["benchmark_name"].str.contains("Nifty 50", na=False)][
        ["snapshot_date","return_1m","return_3m","return_6m","return_1y","return_3y","return_5y"]
    ].rename(columns=lambda c: "bench_"+c if c != "snapshot_date" else c)
    nifty = nifty.groupby("snapshot_date").first().reset_index()
    df    = df_perf.merge(nifty, on="snapshot_date", how="left")
    for p in ["1m","3m","6m","1y","3y","5y"]:
        rc, bc = f"return_{p}", f"bench_return_{p}"
        if rc in df.columns and bc in df.columns:
            df[f"alpha_{p}"] = df[rc] - df[bc]
    return df


# ─────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────
def build_sidebar(df, df_notes):
    with st.sidebar:
        st.markdown("## 📊 APMI PMS Dashboard")
        st.markdown("*Scripbox Internal Tool*")
        st.markdown("---")
        all_strat = sorted(df["strategy_type"].dropna().unique().tolist()) if not df.empty else []
        sel_strat = st.multiselect("Strategy Type", all_strat, default=all_strat)
        all_svc   = sorted(df["service_type"].dropna().unique().tolist()) if not df.empty else []
        sel_svc   = st.multiselect("Service Type", all_svc, default=all_svc)
        period_label = st.selectbox("Return Period", list(RETURN_COLS.values()), index=3)
        period = [k for k,v in RETURN_COLS.items() if v == period_label][0]
        min_aum = st.slider("Min AUM (₹ Cr)", 0, 1000, 0, 50)
        prov_search = st.text_input("Search Provider", "")
        dates = sorted(df["snapshot_date"].dt.strftime("%Y-%m-%d").unique(), reverse=True) if not df.empty else []
        sel_date = st.selectbox("Snapshot Date", dates) if dates else None
        st.markdown("---")
        st.markdown("**🔍 Data Quality Filter**")
        avail_flags = sorted(df["data_quality_flag"].dropna().unique().tolist()) if not df.empty else ["CLEAN"]
        show_flagged = st.toggle("Include flagged rows", value=False)
        if show_flagged:
            sel_flags = st.multiselect("Flags to include", options=avail_flags, default=avail_flags,
                                       format_func=lambda f: FLAG_LABELS.get(f, f))
            if not sel_flags:
                sel_flags = ["CLEAN"]
        else:
            sel_flags = ["CLEAN"]
            st.caption("Showing CLEAN rows only.")
        st.markdown("---")
        if not df_notes.empty:
            with st.expander("📋 Verified Data Notes"):
                for _, n in df_notes.iterrows():
                    snap = f" · {str(n['snapshot_date'])[:10]}" if pd.notna(n.get("snapshot_date")) else ""
                    st.markdown(f"**{n['pms_provider']}** — *{n['ia_name']}*{snap}  \n`{n['note_type']}`  \n{n['explanation']}")
                    st.markdown("---")
        if not df.empty:
            st.markdown(f"*Latest: {df['snapshot_date'].max().strftime('%b %Y')}*")
            st.markdown(f"*{len(df):,} total records*")
    return dict(strategies=sel_strat, service_types=sel_svc, period=period,
                period_label=period_label, min_aum=min_aum, provider_search=prov_search,
                snapshot_date=sel_date, selected_flags=sel_flags, show_flagged=show_flagged)

def apply_filters(df, f):
    if df.empty: return df
    if f.get("snapshot_date"):
        df = df[df["snapshot_date"].dt.strftime("%Y-%m-%d") == f["snapshot_date"]]
    else:
        df = df[df["snapshot_date"] == df["snapshot_date"].max()]
    if f["strategies"]:    df = df[df["strategy_type"].isin(f["strategies"])]
    if f["service_types"]: df = df[df["service_type"].isin(f["service_types"])]
    if f["min_aum"] > 0:   df = df[df["aum_cr"] >= f["min_aum"]]
    if f["provider_search"]:
        df = df[df["pms_provider"].str.lower().str.contains(f["provider_search"].lower(), na=False)]
    if "data_quality_flag" in df.columns and f.get("selected_flags"):
        df = df[df["data_quality_flag"].isin(f["selected_flags"])]
    return df

def explainer(text):
    with st.expander("💡 What does this mean?"):
        st.markdown(f"<div class='info-box'>{text}</div>", unsafe_allow_html=True)

def flag_banner(df, col):
    if "data_quality_flag" not in df.columns: return
    flagged = df[df["data_quality_flag"] != "CLEAN"]
    if flagged.empty: return
    lines = [f"{FLAG_LABELS.get(f,f)}: {c}" for f,c in flagged["data_quality_flag"].value_counts().items()]
    st.markdown(f"<div class='warn-box'>⚠️ Flagged rows visible: {'  |  '.join(lines)}</div>",
                unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────
# VIEW 1: LEADERBOARD
# ─────────────────────────────────────────────────────────
def view_leaderboard(df, filters, df_bench):
    st.markdown("<div class='section-header'>🏆 TWRR Leaderboard</div>", unsafe_allow_html=True)
    explainer("Ranks every PMS Investment Approach by TWRR for the selected period. Q1 = top 25% within strategy peers. Alpha = IA return minus Nifty 50 TRI.")
    if df.empty: st.info("No data."); return
    period, period_label = filters["period"], filters["period_label"]
    df = df.loc[:, ~df.columns.duplicated()]
    if period not in df.columns: st.warning(f"{period} not found."); return
    bench_col, alpha_col = PERIOD_TO_BENCH.get(period, (None, None))
    base = [c for c in ["pms_provider","ia_name","strategy_type","service_type","aum_cr","data_quality_flag"] if c in df.columns]
    extra = [c for c in [bench_col, alpha_col] if c and c in df.columns]
    df_lb = df[base+[period]+extra].dropna(subset=[period]).sort_values(period, ascending=False).copy()
    df_lb.insert(0, "Rank", range(1, len(df_lb)+1))
    def add_q(g):
        g = g.copy()
        try: g["Quartile"] = pd.qcut(g[period], q=4, labels=["Q4","Q3","Q2","Q1"], duplicates="drop")
        except: g["Quartile"] = "N/A"
        return g
    df_lb = df_lb.groupby("strategy_type", group_keys=False).apply(add_q)
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("IAs shown", f"{len(df_lb):,}")
    c2.metric(f"Best ({period_label})", f"{df_lb[period].max():+.2f}%")
    c3.metric(f"Worst ({period_label})", f"{df_lb[period].min():+.2f}%")
    c4.metric(f"Median ({period_label})", f"{df_lb[period].median():+.2f}%")
    if alpha_col and alpha_col in df_lb.columns:
        pct = (df_lb[alpha_col]>0).mean()*100
        bv  = df_lb[bench_col].dropna().iloc[0] if bench_col and bench_col in df_lb.columns and not df_lb[bench_col].dropna().empty else None
        b1,b2 = st.columns(2)
        if bv is not None: b1.metric(f"Nifty 50 TRI ({period_label})", f"{bv:+.2f}%")
        b2.metric("% IAs beat Nifty 50", f"{pct:.1f}%")
    st.markdown("---")
    flag_banner(df_lb, period)
    extreme = df_lb[df_lb[period]<-100]
    if not extreme.empty:
        names = ", ".join(extreme["ia_name"].unique()[:3])
        st.markdown(f"<div class='warn-box'>⚠️ {len(extreme)} row(s) below -100% (e.g. {names}) — verify before presenting.</div>", unsafe_allow_html=True)
    rmap = {"pms_provider":"Provider","ia_name":"Investment Approach","strategy_type":"Strategy",
            "service_type":"Service Type","aum_cr":"AUM (₹ Cr)","data_quality_flag":"Flag",period:f"Return ({period_label})"}
    if bench_col and bench_col in df_lb.columns: rmap[bench_col]=f"Nifty 50 TRI ({period_label})"
    if alpha_col and alpha_col in df_lb.columns: rmap[alpha_col]="Alpha vs Nifty"
    df_d = df_lb.rename(columns=rmap)
    dcols = ["Rank","Provider","Investment Approach","Strategy","AUM (₹ Cr)",f"Return ({period_label})"]
    if f"Nifty 50 TRI ({period_label})" in df_d.columns: dcols.append(f"Nifty 50 TRI ({period_label})")
    if "Alpha vs Nifty" in df_d.columns: dcols.append("Alpha vs Nifty")
    dcols += ["Quartile","Flag"]
    dcols = [c for c in dcols if c in df_d.columns]
    st.dataframe(df_d[dcols], use_container_width=True, height=500)
    dl1,dl2 = st.columns(2)
    with dl1: csv_button(df_d, f"leaderboard_{period_label.replace(' ','_')}.csv", f"⬇️ CSV — {period_label}")
    with dl2: excel_button(df, [c for c in BASE_EXPORT_COLS if c in df.columns],
                           "leaderboard_all_periods.xlsx", df_bench, "📥 Excel — All Periods (9 sheets)")
    fig = px.histogram(df_lb, x=period, nbins=50, color="strategy_type", color_discrete_map=COLORS,
                       labels={period:f"Return ({period_label})","strategy_type":"Strategy"},
                       title=f"Distribution of {period_label} Returns")
    fig.add_vline(x=df_lb[period].median(), line_dash="dash", annotation_text=f"Median: {df_lb[period].median():.2f}%")
    if bench_col and bench_col in df_lb.columns:
        bv2 = df_lb[bench_col].dropna()
        if not bv2.empty:
            fig.add_vline(x=bv2.iloc[0], line_dash="dot", line_color="#f59e0b",
                          annotation_text=f"Nifty 50: {bv2.iloc[0]:.2f}%", annotation_font_color="#f59e0b")
    fig.update_layout(height=350, template="plotly_white")
    st.plotly_chart(fig, use_container_width=True)


# ─────────────────────────────────────────────────────────
# VIEW 2: QUARTILE ANALYSIS
# ─────────────────────────────────────────────────────────
def view_quartile_analysis(df, filters, df_bench):
    st.markdown("<div class='section-header'>📦 Quartile Analysis</div>", unsafe_allow_html=True)
    explainer("Box plots showing return spread within each strategy. Orange dotted line = Nifty 50 TRI.")
    if df.empty: st.info("No data."); return
    period, period_label = filters["period"], filters["period_label"]
    df = df.loc[:, ~df.columns.duplicated()]
    if period not in df.columns: st.warning(f"{period} not available."); return
    df_q = df.dropna(subset=[period])
    bench_col = PERIOD_TO_BENCH.get(period, (None, None))[0]
    bv = None
    if bench_col and bench_col in df_q.columns:
        tmp = df_q[bench_col].dropna()
        if not tmp.empty: bv = tmp.iloc[0]
    c1,c2 = st.columns([3,1])
    with c1:
        fig = px.box(df_q, x="strategy_type", y=period, color="strategy_type", color_discrete_map=COLORS,
                     points="outliers", labels={period:f"Return ({period_label}) %","strategy_type":"Strategy"},
                     title=f"{period_label} Return Distribution by Strategy")
        fig.add_hline(y=0, line_dash="dash", line_color="#6b7280", opacity=0.5)
        if bv is not None:
            fig.add_hline(y=bv, line_dash="dot", line_color="#f59e0b", opacity=0.8,
                          annotation_text=f"Nifty 50 TRI: {bv:.2f}%", annotation_font_color="#f59e0b")
        fig.update_layout(height=450, template="plotly_white", showlegend=False)
        st.plotly_chart(fig, use_container_width=True)
    with c2:
        stats = df_q.groupby("strategy_type")[period].agg(["min","median","mean","max","count"]).round(2)
        stats.columns = ["Min","Median","Mean","Max","Count"]
        st.markdown("**Stats by Strategy**")
        st.dataframe(stats, use_container_width=True)
    dl1,dl2 = st.columns(2)
    with dl1: csv_button(df_q[["pms_provider","ia_name","strategy_type","service_type",period]], "quartile.csv", f"⬇️ CSV — {period_label}")
    with dl2: excel_button(df, [c for c in BASE_EXPORT_COLS if c in df.columns], "quartile_all_periods.xlsx", df_bench, "📥 Excel — All Periods (9 sheets)")


# ─────────────────────────────────────────────────────────
# VIEW 3: AUM vs RETURN
# ─────────────────────────────────────────────────────────
def view_aum_scatter(df, filters, df_bench):
    st.markdown("<div class='section-header'>🔵 AUM vs Return</div>", unsafe_allow_html=True)
    explainer("Each dot = one Investment Approach. Does fund size predict performance?")
    if df.empty: st.info("No data."); return
    period, period_label = filters["period"], filters["period_label"]
    df = df.loc[:, ~df.columns.duplicated()]
    if period not in df.columns: st.warning(f"{period} not available."); return
    df_s = df.dropna(subset=[period,"aum_cr"])
    df_s = df_s[df_s["aum_cr"]>0]
    bench_col = PERIOD_TO_BENCH.get(period, (None, None))[0]
    bv = None
    if bench_col and bench_col in df_s.columns:
        tmp = df_s[bench_col].dropna()
        if not tmp.empty: bv = tmp.iloc[0]
    c1,c2 = st.columns([4,1])
    with c1:
        fig = px.scatter(df_s, x="aum_cr", y=period, color="strategy_type", color_discrete_map=COLORS,
                         size="aum_cr", size_max=20, hover_data={"pms_provider":True,"ia_name":True,"aum_cr":":.0f"},
                         log_x=True, labels={"aum_cr":"AUM (₹ Cr, log scale)",period:f"Return ({period_label}) %"},
                         title=f"AUM vs {period_label} Return")
        fig.add_hline(y=0, line_dash="dash", line_color="#6b7280", opacity=0.4)
        if bv is not None:
            fig.add_hline(y=bv, line_dash="dot", line_color="#f59e0b", opacity=0.8,
                          annotation_text=f"Nifty 50 TRI: {bv:.2f}%", annotation_font_color="#f59e0b")
        fig.update_layout(height=500, template="plotly_white")
        st.plotly_chart(fig, use_container_width=True)
    with c2:
        st.markdown("**Correlation**")
        for strat in df_s["strategy_type"].unique():
            sub = df_s[df_s["strategy_type"]==strat]
            if len(sub)>5:
                corr = sub["aum_cr"].corr(sub[period])
                d = "📈 Positive" if corr>0.1 else ("📉 Negative" if corr<-0.1 else "➡️ Neutral")
                st.markdown(f"**{strat}**: {corr:.3f} {d}")
    dl1,dl2 = st.columns(2)
    with dl1: csv_button(df_s[["pms_provider","ia_name","strategy_type","aum_cr",period]], "aum_scatter.csv", f"⬇️ CSV — {period_label}")
    with dl2: excel_button(df, [c for c in BASE_EXPORT_COLS if c in df.columns], "aum_vs_return_all_periods.xlsx", df_bench, "📥 Excel — All Periods (9 sheets)")


# ─────────────────────────────────────────────────────────
# VIEW 4: TIME-SERIES
# ─────────────────────────────────────────────────────────
def view_time_series(df_all, filters, df_bench):
    st.markdown("<div class='section-header'>📈 Time-Series Evolution</div>", unsafe_allow_html=True)
    explainer("Monthly return evolution for selected IAs. The Excel download gives ALL return periods × ALL months for selected IAs in one file — one sheet per period.")
    if df_all.empty: st.info("No data."); return
    df_all = df_all.loc[:, ~df_all.columns.duplicated()]
    period, period_label = filters["period"], filters["period_label"]
    c1,c2,c3 = st.columns(3)
    with c1:
        strats = st.multiselect("Strategy", df_all["strategy_type"].unique().tolist(), default=["Equity"], key="ts_strat")
    with c2:
        provs = sorted(df_all[df_all["strategy_type"].isin(strats)]["pms_provider"].unique().tolist()) if strats else []
        sel_provs = st.multiselect("Provider", provs, default=provs[:2] if provs else [], key="ts_prov")
    with c3:
        ias = sorted(df_all[df_all["pms_provider"].isin(sel_provs)]["ia_name"].unique().tolist()) if sel_provs else []
        sel_ias = st.multiselect(f"Investment Approach (max 5) — {len(ias)} available",
                                  ias, default=ias[:3] if ias else [], key="ts_ia", max_selections=5)
    show_bench = st.toggle("Overlay Nifty 50 TRI Benchmark", value=True)
    if not sel_ias: st.info("Select at least one Investment Approach above."); return
    if period not in df_all.columns: st.warning(f"{period} not available."); return
    df_ts = df_all[df_all["ia_name"].isin(sel_ias)].dropna(subset=[period]).sort_values("snapshot_date")
    fig = go.Figure()
    for ia in sel_ias:
        sub = df_ts[df_ts["ia_name"]==ia]
        if sub.empty: continue
        fig.add_trace(go.Scatter(x=sub["snapshot_date"], y=sub[period], mode="lines+markers", name=ia,
                                  hovertemplate=f"<b>{ia}</b><br>%{{x|%b %Y}}: %{{y:.2f}}%<extra></extra>",
                                  line=dict(width=2), marker=dict(size=4)))
    if show_bench and not df_bench.empty:
        nifty = df_bench[df_bench["benchmark_name"].str.contains("Nifty 50", na=False)]
        if not nifty.empty and period in nifty.columns:
            nf = nifty.dropna(subset=[period])
            fig.add_trace(go.Scatter(x=nf["snapshot_date"], y=nf[period], mode="lines", name="Nifty 50 TRI",
                                      line=dict(color=COLORS["benchmark"], dash="dash", width=2)))
    fig.add_hline(y=0, line_dash="dot", line_color="#94a3b8", opacity=0.5)
    fig.update_layout(title=f"Monthly {period_label} Return — Selected IAs", xaxis_title="Date",
                      yaxis_title="Return %", height=500, template="plotly_white", hovermode="x unified",
                      legend=dict(orientation="h", yanchor="bottom", y=-0.3))
    st.plotly_chart(fig, use_container_width=True)
    if not df_ts.empty:
        df_hist = df_all[df_all["ia_name"].isin(sel_ias)].copy()
        dl1,dl2 = st.columns(2)
        with dl1:
            csv_button(df_ts[["snapshot_date","pms_provider","ia_name","strategy_type",period,"aum_cr"]],
                       f"timeseries_{period_label.replace(' ','_')}.csv", f"⬇️ CSV — {period_label} only")
        with dl2:
            fn = "timeseries_all_periods_" + "_".join([ia[:10].replace(" ","") for ia in sel_ias[:2]]) + ".xlsx"
            excel_button(df_hist, [c for c in BASE_EXPORT_COLS if c in df_hist.columns], fn, df_bench,
                         "📥 Excel — All Periods × All Months")
        st.caption("Excel: one sheet per return period, each showing all selected fund(s) across all 36 months.")


# ─────────────────────────────────────────────────────────
# VIEW 5: AUM DECOMPOSITION
# ─────────────────────────────────────────────────────────
def view_aum_decomposition(df_aum, filters):
    st.markdown("<div class='section-header'>🧩 AUM Decomposition</div>", unsafe_allow_html=True)
    explainer("Separates AUM growth into Organic (investment returns) vs Net Flows (investor money). ⚠️ Flows are ESTIMATED.")
    if df_aum.empty:
        st.markdown("<div class='warn-box'>Run <code>calculate_aum_decomposition()</code> first.</div>", unsafe_allow_html=True)
        return
    mode = st.radio("View mode", ["📊 Industry Overview","🔎 Fund Deep-Dive"], horizontal=True, key="aum_mode")
    strategies = filters.get("strategies", [])
    if mode == "📊 Industry Overview":
        df_f = df_aum[df_aum["strategy_type"].isin(strategies)] if strategies else df_aum
        monthly = df_f.groupby("snapshot_date").agg(
            organic=("organic_growth","sum"),
            flows=("net_flows","sum"),
            total_aum=("aum_current","sum")
        ).reset_index()
        c1,c2 = st.columns(2)
        with c1:
            fig = make_subplots(specs=[[{"secondary_y":True}]])
            fig.add_trace(go.Bar(x=monthly["snapshot_date"], y=monthly["organic"], name="Organic Growth", marker_color=COLORS["Inflow"], opacity=0.8), secondary_y=False)
            fig.add_trace(go.Bar(x=monthly["snapshot_date"], y=monthly["flows"], name="Est. Net Flows", marker_color=COLORS["benchmark"], opacity=0.8), secondary_y=False)
            fig.add_trace(go.Scatter(x=monthly["snapshot_date"], y=monthly["total_aum"], name="Total AUM", line=dict(color="#1e293b", width=2)), secondary_y=True)
            fig.update_layout(title="Organic Growth vs Net Flows vs Total AUM", barmode="relative", height=420, template="plotly_white")
            st.plotly_chart(fig, use_container_width=True)
        with c2:
            fc = df_f.groupby(["snapshot_date","flow_direction"]).size().reset_index(name="count")
            fig2 = px.bar(fc, x="snapshot_date", y="count", color="flow_direction", color_discrete_map=COLORS,
                          title="IAs by Flow Direction Each Month", barmode="stack")
            fig2.update_layout(height=420, template="plotly_white")
            st.plotly_chart(fig2, use_container_width=True)
        csv_button(df_f, "aum_decomposition.csv")
    else:
        col_a,col_b = st.columns(2)
        with col_a:
            sel_prov = st.selectbox("Provider", sorted(df_aum["pms_provider"].dropna().unique()), key="aum_prov")
        with col_b:
            sel_ia = st.selectbox("Investment Approach", sorted(df_aum[df_aum["pms_provider"]==sel_prov]["ia_name"].dropna().unique()), key="aum_ia")
        df_fund = df_aum[(df_aum["pms_provider"]==sel_prov) & (df_aum["ia_name"]==sel_ia)].sort_values("snapshot_date").copy()
        if df_fund.empty: st.info("No data for this fund."); return
        latest, oldest = df_fund.iloc[-1], df_fund.iloc[0]
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Current AUM", f"₹{latest['aum_current']:,.0f} Cr")
        m2.metric("AUM at Start", f"₹{oldest['aum_current']:,.0f} Cr")
        m3.metric("Total Organic Growth", f"₹{df_fund['organic_growth'].sum():,.0f} Cr")
        tf = df_fund["net_flows"].sum()
        m4.metric("Net Inflows" if tf>0 else "Net Outflows", f"₹{abs(tf):,.0f} Cr")
        fig1 = make_subplots(specs=[[{"secondary_y":True}]])
        fig1.add_trace(go.Bar(x=df_fund["snapshot_date"], y=df_fund["organic_growth"], name="Organic Growth", marker_color=COLORS["Inflow"], opacity=0.8), secondary_y=False)
        fig1.add_trace(go.Bar(x=df_fund["snapshot_date"], y=df_fund["net_flows"], name="Est. Net Flows", marker_color=COLORS["benchmark"], opacity=0.8), secondary_y=False)
        fig1.add_trace(go.Scatter(x=df_fund["snapshot_date"], y=df_fund["aum_current"], name="AUM", line=dict(color="#1e293b", width=2.5)), secondary_y=True)
        fig1.update_layout(title=f"{sel_ia} — Monthly AUM Decomposition", barmode="relative", height=420, template="plotly_white", legend=dict(orientation="h", y=-0.25))
        st.plotly_chart(fig1, use_container_width=True)
        df_fund["cum_org"] = df_fund["organic_growth"].cumsum()
        df_fund["cum_fl"]  = df_fund["net_flows"].cumsum()
        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(x=df_fund["snapshot_date"], y=df_fund["cum_org"], name="Cumulative Organic", line=dict(color=COLORS["Inflow"], width=2)))
        fig2.add_trace(go.Scatter(x=df_fund["snapshot_date"], y=df_fund["cum_fl"], name="Cumulative Flows", line=dict(color=COLORS["benchmark"], width=2, dash="dash")))
        fig2.add_hline(y=0, line_dash="dot", line_color="#94a3b8", opacity=0.5)
        fig2.update_layout(title="Cumulative: Growth from returns vs investor money", xaxis_title="Date", yaxis_title="Cumulative ₹ Cr", height=380, template="plotly_white", legend=dict(orientation="h", y=-0.25))
        st.plotly_chart(fig2, use_container_width=True)
        csv_button(df_fund, f"aum_decomp_{sel_ia[:20].replace(' ','_')}.csv")


# ─────────────────────────────────────────────────────────
# VIEW 6: RISK METRICS
# ─────────────────────────────────────────────────────────
def view_risk_metrics(df_risk, df_perf, filters, df_bench):
    st.markdown("<div class='section-header'>⚖️ Risk-Adjusted Performance</div>", unsafe_allow_html=True)
    explainer("Sharpe = return/risk. Sortino = return/downside risk. Max Drawdown = worst loss from peak. Alpha = excess return vs Nifty 50.")
    if df_risk.empty:
        st.markdown("<div class='warn-box'>Run <code>calculate_risk_metrics()</code> first.</div>", unsafe_allow_html=True); return
    strats = filters.get("strategies", [])
    df_r = df_risk[df_risk["strategy_type"].isin(strats)] if strats else df_risk
    df_perf = df_perf.loc[:, ~df_perf.columns.duplicated()]
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Avg Sharpe", f"{df_r['sharpe_ratio'].mean():.2f}")
    c2.metric("Avg Sortino", f"{df_r['sortino_ratio'].mean():.2f}")
    c3.metric("Avg Max Drawdown", f"{df_r['max_drawdown'].mean():.2f}%")
    pct = (df_r["alpha_vs_nifty50"]>0).mean()*100 if df_r["alpha_vs_nifty50"].notna().any() else 0
    c4.metric("% IAs with +ve Alpha", f"{pct:.1f}%")
    c1b,c2b = st.columns(2)
    with c1b:
        if "return_1y" in df_perf.columns:
            dp = df_r.merge(df_perf[["pms_provider","ia_name","service_type","return_1y"]].dropna(), on=["pms_provider","ia_name","service_type"], how="inner")
            if not dp.empty:
                fig = px.scatter(dp, x="sharpe_ratio", y="return_1y", color="strategy_type", color_discrete_map=COLORS,
                                  hover_data={"pms_provider":True,"ia_name":True},
                                  labels={"sharpe_ratio":"Sharpe Ratio","return_1y":"1Y Return %"}, title="Sharpe vs 1Y Return")
                fig.add_vline(x=dp["sharpe_ratio"].median(), line_dash="dash", line_color="#94a3b8", opacity=0.5)
                fig.add_hline(y=dp["return_1y"].median(), line_dash="dash", line_color="#94a3b8", opacity=0.5)
                fig.update_layout(height=420, template="plotly_white")
                st.plotly_chart(fig, use_container_width=True)
    with c2b:
        fig2 = px.box(df_r, x="strategy_type", y="max_drawdown", color="strategy_type", color_discrete_map=COLORS,
                      title="Max Drawdown by Strategy", labels={"max_drawdown":"Max Drawdown %"})
        fig2.update_layout(height=420, template="plotly_white", showlegend=False)
        st.plotly_chart(fig2, use_container_width=True)
    disp = df_r[["pms_provider","ia_name","strategy_type","sharpe_ratio","sortino_ratio","max_drawdown","alpha_vs_nifty50","beta","std_deviation","months_of_data"]].copy()
    disp.columns = ["Provider","IA","Strategy","Sharpe","Sortino","Max DD %","Alpha","Beta","Std Dev","Months"]
    st.dataframe(disp.sort_values("Sharpe", ascending=False).round(3), use_container_width=True, height=400)
    dl1,dl2 = st.columns(2)
    with dl1: csv_button(disp, "risk_metrics.csv", "⬇️ CSV — Risk metrics")
    with dl2: excel_button(df_perf, [c for c in BASE_EXPORT_COLS if c in df_perf.columns], "risk_context_all_periods.xlsx", df_bench, "📥 Excel — Returns All Periods (9 sheets)")


# ─────────────────────────────────────────────────────────
# VIEW 7: AUM-EFFECT
# ─────────────────────────────────────────────────────────
def view_aum_effect(df_risk, df_perf, filters, df_bench):
    st.markdown("<div class='section-header'>🔬 AUM-Effect Analysis</div>", unsafe_allow_html=True)
    explainer("IAs split into 5 AUM buckets. Q1=smallest, Q5=largest. If Q5>Q1, scale is an advantage.")
    if df_risk.empty or df_perf.empty: st.info("Need both risk metrics and performance data."); return
    period, period_label = filters["period"], filters["period_label"]
    strats = filters.get("strategies", [])
    df_perf = df_perf.loc[:, ~df_perf.columns.duplicated()]
    if period not in df_perf.columns: st.warning(f"{period} not available."); return
    dm = df_risk[["pms_provider","ia_name","service_type","strategy_type","aum_quintile"]].merge(
        df_perf[["pms_provider","ia_name","service_type",period]].dropna(), on=["pms_provider","ia_name","service_type"], how="inner")
    if strats: dm = dm[dm["strategy_type"].isin(strats)]
    dm = dm.dropna(subset=["aum_quintile",period])
    if dm.empty: st.info("Not enough data."); return
    qa = dm.groupby(["strategy_type","aum_quintile"])[period].agg(["mean","median","count"]).reset_index()
    qa.columns = ["Strategy","Quintile","Mean Return","Median Return","Count"]
    qa["Quintile Label"] = qa["Quintile"].astype(str).map({"1":"Q1 (Smallest)","2":"Q2","3":"Q3","4":"Q4","5":"Q5 (Largest)"})
    c1,c2 = st.columns(2)
    with c1:
        fig = px.bar(qa, x="Quintile Label", y="Mean Return", color="Strategy", color_discrete_map=COLORS, barmode="group", title=f"Avg {period_label} Return by AUM Quintile")
        fig.add_hline(y=0, line_dash="dash", line_color="#94a3b8", opacity=0.5)
        fig.update_layout(height=420, template="plotly_white")
        st.plotly_chart(fig, use_container_width=True)
    with c2:
        fig2 = px.box(dm, x="aum_quintile", y=period, color="strategy_type", color_discrete_map=COLORS,
                      title=f"{period_label} Return Spread by AUM Quintile", labels={"aum_quintile":"AUM Quintile",period:"Return %"})
        fig2.add_hline(y=0, line_dash="dash", line_color="#94a3b8", opacity=0.5)
        fig2.update_layout(height=420, template="plotly_white")
        st.plotly_chart(fig2, use_container_width=True)
    st.markdown("#### Interpretation")
    for strat in (strats or qa["Strategy"].unique()):
        sub = qa[qa["Strategy"]==strat]
        q1  = sub[sub["Quintile"]==1]["Mean Return"].values
        q5  = sub[sub["Quintile"]==5]["Mean Return"].values
        if len(q1)>0 and len(q5)>0:
            diff = q5[0]-q1[0]
            d = "scale ADVANTAGE ✅" if diff>1 else ("scale DRAG ⚠️" if diff<-1 else "no clear pattern")
            st.markdown(f"<div class='metric-card'><b>{strat}</b>: Q5={q5[0]:.2f}% vs Q1={q1[0]:.2f}% → diff {diff:+.2f}% → <b>{d}</b></div>", unsafe_allow_html=True)
    dl1,dl2 = st.columns(2)
    with dl1: csv_button(qa, "aum_effect.csv", f"⬇️ CSV — {period_label}")
    with dl2: excel_button(df_perf, [c for c in BASE_EXPORT_COLS if c in df_perf.columns], "aum_effect_all_periods.xlsx", df_bench, "📥 Excel — All Periods (9 sheets)")


# ─────────────────────────────────────────────────────────
# VIEW 8: IA COMPARE
# ─────────────────────────────────────────────────────────
def view_ia_compare(df_perf, df_risk, filters, df_bench):
    st.markdown("<div class='section-header'>🔀 IA Compare</div>", unsafe_allow_html=True)
    explainer("Compare up to 5 IAs side-by-side. Excel download gives full history across all return periods.")
    df_clean = df_perf[df_perf["data_quality_flag"]=="CLEAN"] if "data_quality_flag" in df_perf.columns else df_perf
    latest   = df_clean["snapshot_date"].max()
    df_lat   = df_clean[df_clean["snapshot_date"]==latest].copy()
    df_lat["display_label"] = df_lat["pms_provider"]+" · "+df_lat["ia_name"]
    sel = st.multiselect("Select Investment Approaches (max 5)", sorted(df_lat["display_label"].unique()), max_selections=5)
    if not sel: st.info("Select at least one IA."); return
    df_s     = df_lat[df_lat["display_label"].isin(sel)].copy()
    ia_names = df_s["ia_name"].tolist()
    RCOLS = ["return_1m","return_3m","return_6m","return_1y","return_2y","return_3y","return_5y","return_si"]
    RLBL  = ["1M","3M","6M","1Y","2Y","3Y","5Y","SI"]
    rows  = []
    for _, row in df_s.iterrows():
        r = {"Provider":row["pms_provider"],"IA Name":row["ia_name"],"Strategy":row.get("strategy_type","—"),
             "Service":row.get("service_type","—"),"AUM (₹ Cr)":f"{row['aum_cr']:,.0f}" if pd.notna(row.get("aum_cr")) else "—"}
        for c,l in zip(RCOLS, RLBL):
            v = row.get(c); r[l] = f"{v:+.2f}%" if pd.notna(v) else "—"
        rows.append(r)
    st.markdown("### 📋 Side-by-Side Snapshot")
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    st.markdown("### 📊 Return Comparison by Period")
    bd = []
    for _, row in df_s.iterrows():
        lbl = row["pms_provider"]+" · "+row["ia_name"]
        for c,p in zip(RCOLS, RLBL):
            v = row.get(c)
            if pd.notna(v): bd.append({"IA":lbl,"Period":p,"Return (%)":v})
    if bd:
        df_bar = pd.DataFrame(bd)
        df_bar["Period"] = pd.Categorical(df_bar["Period"], categories=RLBL, ordered=True)
        fig_bar = px.bar(df_bar.sort_values("Period"), x="Period", y="Return (%)", color="IA", barmode="group", color_discrete_sequence=px.colors.qualitative.Set2)
        fig_bar.update_layout(legend=dict(orientation="h", y=-0.3), height=400, template="plotly_white")
        st.plotly_chart(fig_bar, use_container_width=True)
    st.markdown("### 📈 Monthly 1M Return Over Time")
    df_hist = df_clean[df_clean["ia_name"].isin(ia_names)].copy()
    df_hist["snapshot_date"] = pd.to_datetime(df_hist["snapshot_date"])
    fig_ts = go.Figure()
    for ia in ia_names:
        d = df_hist[df_hist["ia_name"]==ia].sort_values("snapshot_date")
        if not d.empty and "return_1m" in d.columns:
            fig_ts.add_trace(go.Scatter(x=d["snapshot_date"], y=d["return_1m"], mode="lines+markers",
                                         name=f"{d['pms_provider'].iloc[0]} · {ia}", line=dict(width=2)))
    fig_ts.update_layout(xaxis_title="Date", yaxis_title="1M Return (%)", legend=dict(orientation="h", y=-0.3), height=420, template="plotly_white")
    st.plotly_chart(fig_ts, use_container_width=True)
    st.markdown("### ⚖️ Risk vs Return")
    if df_risk is not None and not df_risk.empty:
        dr = df_risk[df_risk["ia_name"].isin(ia_names)].copy()
        dr = dr.sort_values("snapshot_date").groupby("ia_name").last().reset_index()
        dr = dr.merge(df_s[["ia_name","return_1y"]], on="ia_name", how="left")
        if not dr.empty and "sharpe_ratio" in dr.columns:
            fig_sc = px.scatter(dr, x="sharpe_ratio", y="return_1y", text="ia_name",
                                 labels={"sharpe_ratio":"Sharpe","return_1y":"1Y Return (%)"}, color_discrete_sequence=["#2563EB"])
            fig_sc.update_traces(textposition="top center", marker=dict(size=12))
            fig_sc.update_layout(height=400, template="plotly_white")
            st.plotly_chart(fig_sc, use_container_width=True)
    dl1,dl2 = st.columns(2)
    with dl1: csv_button(pd.DataFrame(rows), "ia_comparison_snapshot.csv", "⬇️ CSV — Latest snapshot")
    with dl2:
        fn = "ia_compare_all_periods_" + "_".join([ia[:10].replace(" ","") for ia in ia_names[:2]]) + ".xlsx"
        excel_button(df_hist, [c for c in BASE_EXPORT_COLS if c in df_hist.columns], fn, df_bench, "📥 Excel — Full History All Periods")
    st.caption("Excel: one sheet per return period, each showing selected fund(s) across all 36 months.")

# ─────────────────────────────────────────────────────────
# VIEW 9: TURNOVER
# ─────────────────────────────────────────────────────────
def view_turnover(df_turnover, df_perf, filters):
    st.markdown("<div class='section-header'>🔄 Turnover Analysis</div>", unsafe_allow_html=True)
    explainer("Turnover ratio = how actively a manager trades. 100% = full portfolio replaced once a year. High turnover = higher transaction costs = less return to investor.")
    if df_turnover.empty:
        st.markdown("<div class='warn-box'>Run <code>run_turnover_scraper()</code> first.</div>", unsafe_allow_html=True); return

    period, period_label = filters["period"], filters["period_label"]
    strats = filters.get("strategies", [])
    snap   = filters.get("snapshot_date")

    df_t = df_turnover.copy()
    if snap: df_t = df_t[df_t["snapshot_date"].dt.strftime("%Y-%m-%d")==snap]
    else:    df_t = df_t[df_t["snapshot_date"]==df_t["snapshot_date"].max()]
    if strats: df_t = df_t[df_t["strategy_type"].isin(strats)]
    if df_t.empty: st.info("No turnover data for selected filters."); return

    # ── Top metrics ───────────────────────────────────────
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("IAs with data",        f"{len(df_t):,}")
    c2.metric("Avg Turnover",         f"{df_t['turnover_ratio'].mean():.1f}%")
    c3.metric("Median Turnover",      f"{df_t['turnover_ratio'].median():.1f}%")
    c4.metric("High Turnover (>200%)",f"{(df_t['turnover_ratio']>200).sum():,}")
    st.markdown("---")

    # ── Outlier notice ────────────────────────────────────
    outliers = df_t[df_t["turnover_ratio"]>300]
    if not outliers.empty:
        names = ", ".join(outliers["ia_name"].unique()[:3])
        st.markdown(f"<div class='warn-box'>⚠️ {len(outliers)} IA(s) above 300% (e.g. {names}). Typically arbitrage strategies — excluded from the distribution chart below but shown in the leaderboard table.</div>", unsafe_allow_html=True)

    # ── Mode toggle ───────────────────────────────────────
    mode = st.radio("View mode", ["📊 Strategy Overview", "🔎 Fund Deep-Dive"], horizontal=True, key="turnover_mode")

    if mode == "📊 Strategy Overview":
        df_normal = df_t[df_t["turnover_ratio"] <= 300]

        cl,cr = st.columns(2)
        with cl:
            fig = px.box(
                df_normal.dropna(subset=["turnover_ratio"]),
                x="strategy_type", y="turnover_ratio", color="strategy_type",
                color_discrete_map=COLORS, points="outliers",
                title="Turnover Distribution by Strategy (capped at 300%)",
                labels={"turnover_ratio":"Turnover (%)","strategy_type":"Strategy"}
            )
            fig.update_layout(height=400, template="plotly_white", showlegend=False)
            st.plotly_chart(fig, use_container_width=True)
        with cr:
            t20 = df_t.nlargest(20,"turnover_ratio")[["pms_provider","ia_name","strategy_type","turnover_ratio"]]
            st.markdown("**Top 20 Highest Turnover IAs**")
            st.dataframe(t20.round(1), use_container_width=True, hide_index=True)

        # Turnover vs Return scatter
        if not df_perf.empty and period in df_perf.columns:
            st.markdown("### Turnover vs Return")
            snap_str = snap if snap else df_t["snapshot_date"].max().strftime("%Y-%m-%d")
            dp = df_perf[df_perf["snapshot_date"].dt.strftime("%Y-%m-%d")==snap_str]
            ds = df_normal.merge(
                dp[["pms_provider","ia_name","service_type",period]].dropna(),
                on=["pms_provider","ia_name","service_type"], how="inner"
            )
            if not ds.empty:
                fig_sc = px.scatter(
                    ds.dropna(subset=["turnover_ratio",period]),
                    x="turnover_ratio", y=period,
                    color="strategy_type", color_discrete_map=COLORS,
                    hover_data={"pms_provider":True,"ia_name":True},
                    labels={"turnover_ratio":"Turnover (%)","return_1y":f"Return ({period_label}) %"},
                    title=f"Turnover vs {period_label} Return (funds ≤300% turnover)"
                )
                fig_sc.add_hline(y=0, line_dash="dash", line_color="#94a3b8", opacity=0.5)
                fig_sc.update_layout(height=450, template="plotly_white")
                st.plotly_chart(fig_sc, use_container_width=True)

        csv_button(df_t, "turnover.csv")

    else:
        # ── Fund Deep-Dive ────────────────────────────────
        col_a, col_b = st.columns(2)
        with col_a:
            sel_prov = st.selectbox("Provider", sorted(df_turnover["pms_provider"].dropna().unique()), key="turn_prov")
        with col_b:
            ia_opts = sorted(df_turnover[df_turnover["pms_provider"]==sel_prov]["ia_name"].dropna().unique())
            sel_ia  = st.selectbox("Investment Approach", ia_opts, key="turn_ia")

        df_fund = df_turnover[
            (df_turnover["pms_provider"]==sel_prov) &
            (df_turnover["ia_name"]==sel_ia)
        ].sort_values("snapshot_date").copy()

        if df_fund.empty: st.info("No turnover data for this fund."); return

        # Summary metrics
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Latest Turnover",  f"{df_fund.iloc[-1]['turnover_ratio']:.1f}%")
        m2.metric("Average Turnover", f"{df_fund['turnover_ratio'].mean():.1f}%")
        m3.metric("Peak Turnover",    f"{df_fund['turnover_ratio'].max():.1f}%")
        trend = df_fund['turnover_ratio'].iloc[-1] - df_fund['turnover_ratio'].iloc[0]
        m4.metric("Trend (first→last)", f"{trend:+.1f}%",
                  delta_color="inverse")  # rising turnover = bad = red

        # Monthly turnover history
        fig1 = go.Figure()
        fig1.add_trace(go.Bar(
            x=df_fund["snapshot_date"], y=df_fund["turnover_ratio"],
            name="Monthly Turnover", marker_color="#2563eb", opacity=0.8
        ))
        fig1.add_hline(
            y=df_fund["turnover_ratio"].mean(),
            line_dash="dash", line_color="#f59e0b",
            annotation_text=f"Avg: {df_fund['turnover_ratio'].mean():.1f}%",
            annotation_font_color="#f59e0b"
        )
        fig1.update_layout(
            title=f"{sel_ia} — Monthly Turnover History",
            xaxis_title="Date", yaxis_title="Turnover (%)",
            height=400, template="plotly_white"
        )
        st.plotly_chart(fig1, use_container_width=True)

        # Overlay with return if available
        if not df_perf.empty and "return_1m" in df_perf.columns:
            df_ret = df_perf[
                (df_perf["pms_provider"]==sel_prov) &
                (df_perf["ia_name"]==sel_ia)
            ][["snapshot_date","return_1m"]].sort_values("snapshot_date")

            if not df_ret.empty:
                merged = df_fund.merge(df_ret, on="snapshot_date", how="inner")
                if not merged.empty:
                    fig2 = make_subplots(specs=[[{"secondary_y":True}]])
                    fig2.add_trace(go.Bar(
                        x=merged["snapshot_date"], y=merged["turnover_ratio"],
                        name="Turnover (%)", marker_color="#2563eb", opacity=0.5
                    ), secondary_y=False)
                    fig2.add_trace(go.Scatter(
                        x=merged["snapshot_date"], y=merged["return_1m"],
                        name="1M Return (%)", line=dict(color="#059669", width=2)
                    ), secondary_y=True)
                    fig2.add_hline(y=0, line_dash="dot", line_color="#94a3b8", opacity=0.4, secondary_y=True)
                    fig2.update_layout(
                        title=f"{sel_ia} — Turnover vs 1M Return",
                        height=400, template="plotly_white",
                        legend=dict(orientation="h", y=-0.25)
                    )
                    fig2.update_yaxes(title_text="Turnover (%)", secondary_y=False)
                    fig2.update_yaxes(title_text="1M Return (%)", secondary_y=True)
                    st.plotly_chart(fig2, use_container_width=True)

        # Detail table
        with st.expander("📋 Monthly Detail Table"):
            st.dataframe(
                df_fund[["snapshot_date","turnover_ratio"]].rename(
                    columns={"snapshot_date":"Date","turnover_ratio":"Turnover (%)"}
                ).sort_values("Date", ascending=False).round(2),
                use_container_width=True, hide_index=True
            )

        csv_button(df_fund[["snapshot_date","pms_provider","ia_name","turnover_ratio"]],
                   f"turnover_{sel_ia[:20].replace(' ','_')}.csv")



# ─────────────────────────────────────────────────────────
# VIEW 10: PROVIDER COMPARE
# ─────────────────────────────────────────────────────────
def view_provider_compare(df_perf, df_risk, df_turnover, filters, df_bench):
    st.markdown("<div class='section-header'>🏢 Provider Compare</div>", unsafe_allow_html=True)
    explainer("Compare PMS firms at firm level. Metrics aggregated across all CLEAN IAs per provider.")
    df_clean = df_perf[df_perf["data_quality_flag"]=="CLEAN"].copy() if "data_quality_flag" in df_perf.columns else df_perf.copy()
    period, period_label = filters["period"], filters["period_label"]
    strats = filters.get("strategies", [])
    snap   = filters.get("snapshot_date")
    df_snap = df_clean[df_clean["snapshot_date"].dt.strftime("%Y-%m-%d")==snap] if snap else df_clean[df_clean["snapshot_date"]==df_clean["snapshot_date"].max()]
    if strats: df_snap = df_snap[df_snap["strategy_type"].isin(strats)]
    if df_snap.empty or period not in df_snap.columns: st.info("No data."); return
    agg = df_snap.groupby("pms_provider").agg(
        ia_count=("ia_name","nunique"), total_aum=("aum_cr","sum"),
        median_return=(period,"median"), mean_return=(period,"mean"),
        best_return=(period,"max"), worst_return=(period,"min"),
        pct_positive=(period, lambda x: (x>0).mean()*100)
    ).round(2).reset_index()
    if not df_risk.empty:
        ra  = df_risk.sort_values("snapshot_date").groupby("pms_provider").last().reset_index()
        agg = agg.merge(ra[["pms_provider","sharpe_ratio","sortino_ratio","max_drawdown","alpha_vs_nifty50"]].round(3), on="pms_provider", how="left")
    if not df_turnover.empty:
        dt = df_turnover.copy()
        if snap: dt = dt[dt["snapshot_date"].dt.strftime("%Y-%m-%d")==snap]
        else:    dt = dt[dt["snapshot_date"]==dt["snapshot_date"].max()]
        if strats: dt = dt[dt["strategy_type"].isin(strats)]
        if not dt.empty:
            ta = dt.groupby("pms_provider")["turnover_ratio"].median().round(1).reset_index()
            ta.columns = ["pms_provider","median_turnover"]
            agg = agg.merge(ta, on="pms_provider", how="left")
    st.markdown("### 📋 Provider Leaderboard")
    rmap = {"pms_provider":"Provider","ia_count":"# IAs","total_aum":"Total AUM (₹ Cr)",
            "median_return":f"Median Return ({period_label})","mean_return":f"Mean Return ({period_label})",
            "best_return":f"Best IA ({period_label})","worst_return":f"Worst IA ({period_label})",
            "pct_positive":"% IAs Positive","sharpe_ratio":"Avg Sharpe","max_drawdown":"Avg Max DD %",
            "alpha_vs_nifty50":"Avg Alpha","median_turnover":"Median Turnover %"}
    df_lb = agg.rename(columns=rmap).sort_values(f"Median Return ({period_label})", ascending=False).reset_index(drop=True)
    df_lb.insert(0, "Rank", range(1, len(df_lb)+1))
    dcols = [c for c in rmap.values() if c in df_lb.columns]
    st.dataframe(df_lb[["Rank"]+dcols], use_container_width=True, height=500, hide_index=True)
    dl1,dl2 = st.columns(2)
    with dl1: csv_button(df_lb, f"provider_leaderboard_{period_label.replace(' ','_')}.csv", f"⬇️ CSV — {period_label}")
    with dl2: excel_button(df_snap, [c for c in BASE_EXPORT_COLS if c in df_snap.columns], "provider_all_periods.xlsx", df_bench, "📥 Excel — All Periods (9 sheets)")
    st.markdown("---")
    st.markdown("### 🔍 Head-to-Head Comparison")
    sel_provs = st.multiselect("Select providers (max 6)", sorted(agg["pms_provider"].unique()), default=[], max_selections=6, key="prov_h2h")
    if not sel_provs: st.info("Select at least two providers."); return
    df_sel  = agg[agg["pms_provider"].isin(sel_provs)].copy()
    scols   = {k:v for k,v in rmap.items() if k in df_sel.columns}
    df_sum  = df_sel[list(scols)].rename(columns=scols)
    st.dataframe(df_sum.set_index("Provider").T, use_container_width=True)
    st.markdown("#### 📦 Return Distribution per Provider")
    df_ia_sel = df_snap[df_snap["pms_provider"].isin(sel_provs)].dropna(subset=[period])
    if not df_ia_sel.empty:
        fig_b = px.box(df_ia_sel, x="pms_provider", y=period, color="pms_provider", points="outliers",
                       hover_data={"ia_name":True}, labels={period:f"Return ({period_label}) %","pms_provider":"Provider"},
                       title=f"{period_label} Return — All IAs per Provider")
        fig_b.add_hline(y=0, line_dash="dash", line_color="#6b7280", opacity=0.4)
        bc = PERIOD_TO_BENCH.get(period,(None,None))[0]
        if bc and bc in df_snap.columns:
            bv = df_snap[bc].dropna()
            if not bv.empty: fig_b.add_hline(y=bv.iloc[0], line_dash="dot", line_color="#f59e0b", opacity=0.8, annotation_text=f"Nifty 50 TRI: {bv.iloc[0]:.2f}%", annotation_font_color="#f59e0b")
        fig_b.update_layout(height=450, template="plotly_white", showlegend=False, xaxis_tickangle=-20)
        st.plotly_chart(fig_b, use_container_width=True)
    st.markdown("#### 📊 Key Metrics Comparison")
    mopts = {f"Median Return ({period_label})":"median_return","Total AUM (₹ Cr)":"total_aum",
             "% IAs Positive":"pct_positive","Avg Sharpe":"sharpe_ratio","Avg Max DD %":"max_drawdown",
             "Avg Alpha":"alpha_vs_nifty50","Median Turnover %":"median_turnover"}
    avail      = {k:v for k,v in mopts.items() if v in df_sel.columns}
    chosen_lbl = st.selectbox("Metric", list(avail), key="prov_metric")
    chosen_col = avail[chosen_lbl]
    df_bbar    = df_sel[["pms_provider",chosen_col]].dropna().sort_values(chosen_col, ascending=False)
    fig_bb = px.bar(df_bbar, x="pms_provider", y=chosen_col, color="pms_provider",
                    labels={chosen_col:chosen_lbl,"pms_provider":"Provider"}, title=f"{chosen_lbl}")
    if chosen_col not in ["total_aum","ia_count","median_turnover"]:
        fig_bb.add_hline(y=0, line_dash="dash", line_color="#6b7280", opacity=0.4)
    fig_bb.update_layout(height=380, template="plotly_white", showlegend=False, xaxis_tickangle=-20)
    st.plotly_chart(fig_bb, use_container_width=True)
    st.markdown("#### 📈 Median Monthly Return Over Time")
    dac = df_clean.copy()
    if strats: dac = dac[dac["strategy_type"].isin(strats)]
    dtp = dac[dac["pms_provider"].isin(sel_provs)]
    if "return_1m" in dtp.columns and not dtp.empty:
        tsa = dtp.groupby(["snapshot_date","pms_provider"])["return_1m"].median().reset_index()
        fig_ts = px.line(tsa, x="snapshot_date", y="return_1m", color="pms_provider",
                         labels={"return_1m":"Median 1M Return (%)","snapshot_date":"Date","pms_provider":"Provider"},
                         title="Median 1M Return per Provider — All Months")
        fig_ts.add_hline(y=0, line_dash="dot", line_color="#94a3b8", opacity=0.5)
        fig_ts.update_layout(height=420, template="plotly_white", hovermode="x unified", legend=dict(orientation="h", y=-0.3))
        st.plotly_chart(fig_ts, use_container_width=True)
    st.markdown("#### 💰 Total AUM Over Time")
    dat = dac[dac["pms_provider"].isin(sel_provs)]
    if "aum_cr" in dat.columns and not dat.empty:
        aa = dat.groupby(["snapshot_date","pms_provider"])["aum_cr"].sum().reset_index()
        fig_a = px.line(aa, x="snapshot_date", y="aum_cr", color="pms_provider",
                        labels={"aum_cr":"Total AUM (₹ Cr)","snapshot_date":"Date","pms_provider":"Provider"},
                        title="Total AUM per Provider — All Months")
        fig_a.update_layout(height=400, template="plotly_white", hovermode="x unified", legend=dict(orientation="h", y=-0.3))
        st.plotly_chart(fig_a, use_container_width=True)
    dl1,dl2 = st.columns(2)
    with dl1: csv_button(df_sum, "provider_comparison.csv", "⬇️ CSV — Provider summary")
    with dl2:
        src = df_ia_sel if not df_ia_sel.empty else df_snap
        excel_button(src, [c for c in BASE_EXPORT_COLS if c in src.columns], "provider_compare_all_periods.xlsx", df_bench, "📥 Excel — Selected Providers All Periods")


# ─────────────────────────────────────────────────────────
# VIEW 11: DATA NOTES
# ─────────────────────────────────────────────────────────
def view_data_notes(df_notes, df_perf):
    st.markdown("<div class='section-header'>📋 Data Source & Notes</div>", unsafe_allow_html=True)

    st.markdown("### 🔌 Data Collection")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("""
**Main Performance Scraper**
- **Endpoint:** `apmiindia.org/welcomeiaperformance.htm?action=loadIAReport`
- **Parameters:** `strategyname` × `servicetype` × `reportdate`
- **Loops:** 4 strategies × 2 service types × 36 months = **288 API calls**
- **What it captures:** TWRR returns (1M, 3M, 6M, 1Y, 2Y, 3Y, 4Y, 5Y, SI) + AUM per IA per month
- **Auth:** None — public SEBI-mandated disclosure
""")
    with col2:
        st.markdown("""
**Turnover Scraper**
- **Endpoint:** `apmiindia.org/IATurnoverReportUtility.htm?action=loadIATurnoverReport`
- **Parameters:** `stretegyname` *(APMI typo — must keep or requests fail)* × `reportdate`
- **Date format:** `YYYY-M-DD` with no leading zero on month (e.g. `2026-3-31`)
- **What it captures:** Portfolio turnover ratio per IA per month
- **Rows scraped:** ~50,128
""")

    st.markdown("---")
    st.markdown("### 📦 Database Coverage")
    if not df_perf.empty:
        c1,c2,c3,c4,c5 = st.columns(5)
        c1.metric("Total Records", f"{len(df_perf):,}")
        c2.metric("Investment Approaches", f"{df_perf['ia_name'].nunique():,}")
        c3.metric("PMS Providers", f"{df_perf['pms_provider'].nunique():,}")
        date_min = df_perf["snapshot_date"].min().strftime("%b %Y")
        date_max = df_perf["snapshot_date"].max().strftime("%b %Y")
        c4.metric("Date Range", f"{date_min} – {date_max}")
        c5.metric("Monthly Snapshots", f"{df_perf['snapshot_date'].nunique()}")
        if "data_quality_flag" in df_perf.columns:
            st.markdown("**Records by Data Quality Flag**")
            flag_counts = df_perf["data_quality_flag"].value_counts().reset_index()
            flag_counts.columns = ["Flag", "Count"]
            flag_counts["Label"] = flag_counts["Flag"].map(FLAG_LABELS).fillna(flag_counts["Flag"])
            flag_counts["% of Total"] = (flag_counts["Count"] / len(df_perf) * 100).round(1)
            st.dataframe(flag_counts[["Label","Count","% of Total"]], use_container_width=False, hide_index=True)

    st.markdown("---")
    st.markdown("### 🔍 Known Data Issues")
    st.caption(f"{len(df_notes):,} verified entries in the data_notes table")

    if df_notes.empty:
        st.info("No entries in data_notes table.")
    else:
        NOTE_COLORS = {
            "APMI_REPORTING_ERROR":   ("#fef2f2", "#991b1b", "🔴"),
            "APMI_REPORTED_ANOMALY":  ("#fef2f2", "#991b1b", "🔴"),
            "SUSPECTED_DATA_ERROR":   ("#fef2f2", "#991b1b", "🔴"),
            "UNUSUAL_MOVE":           ("#fffbeb", "#92400e", "🟡"),
            "WOUND_DOWN":             ("#fffbeb", "#92400e", "🟡"),
            "STRUCTURED_PRODUCT":     ("#fffbeb", "#92400e", "🟡"),
            "VERIFIED_EVENT":         ("#f0fdf4", "#166534", "🟢"),
            "CLEAN_EXPLANATION":      ("#f0fdf4", "#166534", "🟢"),
        }
        DEFAULT_COLOR = ("#f8fafc", "#334155", "⚪")

        fc1,fc2 = st.columns([2,2])
        with fc1:
            all_types = sorted(df_notes["note_type"].dropna().unique().tolist())
            sel_types = st.multiselect("Filter by Note Type", all_types, default=all_types, key="notes_type_filter")
        with fc2:
            prov_search = st.text_input("Search Provider / IA", "", key="notes_prov_search")

        df_show = df_notes.copy()
        if sel_types:
            df_show = df_show[df_show["note_type"].isin(sel_types)]
        if prov_search:
            mask = (df_show["pms_provider"].str.lower().str.contains(prov_search.lower(), na=False) |
                    df_show["ia_name"].str.lower().str.contains(prov_search.lower(), na=False))
            df_show = df_show[mask]

        st.caption(f"Showing {len(df_show):,} of {len(df_notes):,} notes")

        for _, row in df_show.iterrows():
            note_type = str(row.get("note_type", "")).upper()
            bg, text_color, dot = NOTE_COLORS.get(note_type, DEFAULT_COLOR)
            snap = f" · {str(row['snapshot_date'])[:10]}" if pd.notna(row.get("snapshot_date")) else ""
            st.markdown(f"""
<div style="background:{bg};border-radius:8px;padding:12px 16px;margin:6px 0;border-left:4px solid {text_color}">
  <div style="font-size:13px;font-weight:600;color:{text_color};margin-bottom:2px">
    {dot} {row['pms_provider']} — {row['ia_name']}{snap}
  </div>
  <div style="font-size:11px;color:{text_color};opacity:0.8;margin-bottom:6px">
    <code style="background:transparent;font-size:11px">{row['note_type']}</code>
  </div>
  <div style="font-size:13px;color:#1e293b">{row['explanation']}</div>
</div>
""", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### 📜 Regulatory Basis")
    st.markdown("""
<div class='info-box'>
<b>SEBI Circular: SEBI/HO/IMD/IMD-II DOF3/P/CIR/2022/0169 — December 2022</b><br><br>
All SEBI-registered Portfolio Management Service (PMS) providers are required to report monthly
Time-Weighted Rate of Return (TWRR) performance data to APMI (Association of Portfolio Managers in India).
APMI publishes this data publicly on <code>apmiindia.org</code>. This dashboard scrapes and presents
that public disclosure data. <b>All returns are TWRR as mandated.</b>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────
def main():
    df_perf     = load_performance_with_benchmarks()
    df_bench    = load_benchmarks()
    df_risk     = load_risk()
    df_aum      = load_aum_decomp()
    df_turnover = load_turnover()
    df_notes    = load_data_notes()

    if df_perf.empty:
        st.warning("⚠️ Database not found. Make sure apmi_pms.db is in the working directory.")
        return

    filters     = build_sidebar(df_perf, df_notes)
    df_filtered = apply_filters(df_perf.copy(), filters)

    st.markdown("""<div style="padding:24px 0 8px 0">
        <h1 style="color:#1e40af;margin:0;font-size:32px">📊 APMI PMS Performance Dashboard</h1>
        <p style="color:#64748b;margin:4px 0 0 0;font-size:15px">SEBI-mandated TWRR · All registered PMS Investment Approaches · Scripbox</p>
    </div>""", unsafe_allow_html=True)

    if not df_filtered.empty:
        k1,k2,k3,k4,k5 = st.columns(5)
        k1.metric("Investment Approaches", f"{df_filtered['ia_name'].nunique():,}")
        k2.metric("PMS Providers", f"{df_filtered['pms_provider'].nunique():,}")
        k3.metric("Total AUM", f"₹{df_filtered['aum_cr'].sum():,.0f} Cr")
        avg_1y = df_filtered["return_1y"].mean() if "return_1y" in df_filtered.columns else None
        k4.metric("Avg 1Y Return", f"{avg_1y:+.2f}%" if pd.notna(avg_1y) else "—")
        k5.metric("Latest Data", df_perf["snapshot_date"].max().strftime("%b %Y"))

    if filters.get("show_flagged"):
        st.markdown("<div class='warn-box'>⚠️ <b>Flagged data enabled.</b> Some rows may contain APMI errors or wound-down strategies.</div>", unsafe_allow_html=True)

    st.markdown("---")

    tab1,tab2,tab3,tab4,tab5,tab6,tab7,tab8,tab9,tab10,tab11 = st.tabs([
        "🏆 Leaderboard","📦 Quartile Analysis","🔵 AUM vs Return",
        "📈 Time-Series","🧩 AUM Decomposition","⚖️ Risk Metrics",
        "🔬 AUM-Effect","🔀 IA Compare","🔄 Turnover","🏢 Provider Compare",
        "📋 Data Notes"
    ])

    with tab1:  view_leaderboard(df_filtered, filters, df_bench)
    with tab2:  view_quartile_analysis(df_filtered, filters, df_bench)
    with tab3:  view_aum_scatter(df_filtered, filters, df_bench)
    with tab4:  view_time_series(df_perf, filters, df_bench)
    with tab5:  view_aum_decomposition(df_aum, filters)
    with tab6:  view_risk_metrics(df_risk, df_filtered, filters, df_bench)
    with tab7:  view_aum_effect(df_risk, df_filtered, filters, df_bench)
    with tab8:  view_ia_compare(df_perf, df_risk, filters, df_bench)
    with tab9:  view_turnover(df_turnover, df_perf, filters)
    with tab10: view_provider_compare(df_perf, df_risk, df_turnover, filters, df_bench)
    with tab11: view_data_notes(df_notes, df_perf)

    st.markdown("---")
    st.markdown("""<div style="text-align:center;color:#94a3b8;font-size:12px;padding:8px">
        Data: APMI · TWRR per SEBI Dec 2022 circular · AUM decomposition = estimated · Not investment advice · Scripbox
    </div>""", unsafe_allow_html=True)


if __name__ == "__main__":
    main()
